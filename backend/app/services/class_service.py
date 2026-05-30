"""Class service"""
from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, List

from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessException
from app.core.security import get_password_hash
from app.models.entities import (
    Announcement,
    AnnouncementClass,
    AnnouncementRead,
    Class,
    Course,
    Project,
    ProjectLike,
    QuizAttempt,
    StudentClassEnrollment,
    StudentProgress,
    TaskCompletion,
    User,
)

logger = logging.getLogger(__name__)

DEFAULT_STUDENT_PASSWORD = "123456"


def _now_iso(value: datetime | None) -> str:
    return value.isoformat() if value else ""


def _owned_class_query(db: Session, teacher_id: str):
    return db.query(Class).join(Course, Course.id == Class.course_id).filter(Course.created_by == teacher_id)


def list_classes(db: Session, teacher_id: str, course_id: int | None = None, keyword: str | None = None):
    query = _owned_class_query(db, teacher_id)
    if course_id:
        query = query.filter(Class.course_id == course_id)
    if keyword:
        query = query.filter(Class.name.like(f"%{keyword}%"))
    classes = query.order_by(Class.created_at.desc(), Class.id.desc()).all()
    result = []
    for cls in classes:
        student_count = db.query(StudentClassEnrollment).filter(
            StudentClassEnrollment.class_id == cls.id).count()
        result.append({
            "id": cls.id,
            "name": cls.name,
            "course_id": cls.course_id,
            "course_name": cls.course.name if cls.course else "",
            "student_count": student_count,
            "created_at": _now_iso(cls.created_at),
        })
    return result


def create_class(db: Session, name: str, course_id: int, teacher_id: str):
    course = db.query(Course).filter(Course.id == course_id, Course.created_by == teacher_id).first()
    if not course:
        raise BusinessException(404, "课程不存在")
    existing = db.query(Class).filter(
        Class.name == name, Class.course_id == course_id).first()
    if existing:
        raise BusinessException(400, "班级已存在")
    cls = Class(name=name, course_id=course_id)
    try:
        db.add(cls)
        db.commit()
        db.refresh(cls)
        logger.info(f"班级创建: id={cls.id}, name={name}, course_id={course_id}")
    except SQLAlchemyError:
        db.rollback()
        raise BusinessException(500, "创建班级失败")
    return cls


def _delete_student_data(db: Session, student_id: str):
    project_ids = [row.id for row in db.query(Project.id).filter(
        Project.author_id == student_id).all()]
    if project_ids:
        db.query(ProjectLike).filter(ProjectLike.project_id.in_(
            project_ids)).delete(synchronize_session=False)
    db.query(ProjectLike).filter(ProjectLike.user_id ==
                                 student_id).delete(synchronize_session=False)
    db.query(Project).filter(Project.author_id ==
                             student_id).delete(synchronize_session=False)
    db.query(QuizAttempt).filter(QuizAttempt.user_id ==
                                 student_id).delete(synchronize_session=False)
    db.query(StudentProgress).filter(StudentProgress.user_id ==
                                     student_id).delete(synchronize_session=False)
    db.query(AnnouncementRead).filter(AnnouncementRead.user_id ==
                                      student_id).delete(synchronize_session=False)
    db.query(TaskCompletion).filter(TaskCompletion.user_id ==
                                    student_id).delete(synchronize_session=False)
    db.query(StudentClassEnrollment).filter(
        StudentClassEnrollment.user_id == student_id).delete(synchronize_session=False)
    db.query(User).filter(User.id == student_id, User.role ==
                          "student").delete(synchronize_session=False)


def delete_class(db: Session, class_id: int, teacher_id: str):
    cls = _owned_class_query(db, teacher_id).filter(Class.id == class_id).first()
    if not cls:
        return None
    try:
        student_count = db.query(StudentClassEnrollment).filter(
            StudentClassEnrollment.class_id == class_id).count()
        if student_count > 0:
            raise BusinessException(400, "该班级仍有学生，无法删除，请先在学生管理中移除学生")

        class_announcement_ids = [
            row.announcement_id
            for row in db.query(AnnouncementClass.announcement_id).filter(AnnouncementClass.class_id == class_id).all()
        ]
        if class_announcement_ids:
            db.query(AnnouncementRead).filter(
                AnnouncementRead.announcement_id.in_(class_announcement_ids)
            ).delete(synchronize_session=False)
            db.query(TaskCompletion).filter(
                TaskCompletion.announcement_id.in_(class_announcement_ids)
            ).delete(synchronize_session=False)
            db.query(AnnouncementClass).filter(AnnouncementClass.class_id == class_id).delete(synchronize_session=False)
            db.query(Announcement).filter(Announcement.id.in_(class_announcement_ids)).delete(synchronize_session=False)
        db.delete(cls)
        db.commit()
        logger.info(f"班级删除: class_id={class_id}, name={cls.name}")
    except SQLAlchemyError:
        db.rollback()
        raise BusinessException(500, "删除班级失败")
    return cls


def list_class_students(db: Session, class_id: int, teacher_id: str):
    cls = _owned_class_query(db, teacher_id).filter(Class.id == class_id).first()
    if not cls:
        return None
    enrollments = (
        db.query(StudentClassEnrollment)
        .filter(StudentClassEnrollment.class_id == class_id)
        .order_by(StudentClassEnrollment.enrolled_at.desc())
        .all()
    )
    result = []
    for enrollment in enrollments:
        student = db.query(User).filter(User.id == enrollment.user_id).first()
        if not student:
            continue
        result.append({
            "id": student.id,
            "name": student.name,
            "major": student.major or "",
            "enrolled_at": _now_iso(enrollment.enrolled_at),
        })
    return result


def enroll_student(db: Session, class_id: int, student_id: str, teacher_id: str, name: str = ""):
    # 查找班级是否存在
    cls = _owned_class_query(db, teacher_id).filter(Class.id == class_id).first()
    if not cls:
        raise BusinessException(404, "班级不存在")

    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        if not name:
            raise BusinessException(404, "学生不存在，请提供姓名以自动创建账号")
        # 自动创建学生账号，默认密码含字母+数字满足复杂度
        DEFAULT_PASSWORD = "a123456"
        student = User(
            id=student_id,
            name=name,
            hashed_password=get_password_hash(DEFAULT_PASSWORD),
            role="student",
            major="",
            needs_password_change=True,
        )
        db.add(student)
        db.flush()
        logger.info(
            f"自动创建学生账号: student_id={student_id}, name={name}, class_id={class_id}")

    if student.role != "student":
        raise BusinessException(400, "仅可添加学生账号")

    existing = db.query(StudentClassEnrollment).filter(
        StudentClassEnrollment.class_id == class_id,
        StudentClassEnrollment.user_id == student_id,
    ).first()
    if existing:
        return existing, "已存在"
    enrollment = StudentClassEnrollment(user_id=student_id, class_id=class_id)
    try:
        db.add(enrollment)
        db.commit()
        db.refresh(enrollment)
    except SQLAlchemyError:
        db.rollback()
        raise BusinessException(500, "添加学生失败")
    return enrollment, "created"


def remove_student(db: Session, class_id: int, student_id: str, teacher_id: str):
    cls = _owned_class_query(db, teacher_id).filter(Class.id == class_id).first()
    if not cls:
        return None
    enrollment = db.query(StudentClassEnrollment).filter(
        StudentClassEnrollment.class_id == class_id,
        StudentClassEnrollment.user_id == student_id,
    ).first()
    if not enrollment:
        return None
    try:
        db.delete(enrollment)
        db.commit()
    except SQLAlchemyError:
        db.rollback()
        raise BusinessException(500, "移除学生失败")
    return enrollment


def import_students_from_excel(db: Session, file_bytes: bytes, teacher_id: str, class_id: int | None = None):
    """从 Excel 导入学生，自动适配教务系统导出格式。

    适配逻辑：
    1. 扫描前 15 行，找到包含「学号」的行作为表头行
    2. 在表头行中定位「学号」「姓名」所在列
    3. 「姓名」列的下一列视为专业列（如无则留空）
    4. 表头行以下、学号非空的行作为数据行

    传入 class_id 时，自动将导入的学生注册到该班级。
    """
    # 如果指定了班级，先验证班级存在
    target_class = None
    if class_id is not None:
        target_class = _owned_class_query(db, teacher_id).filter(Class.id == class_id).first()
        if not target_class:
            raise BusinessException(404, "班级不存在")

    try:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        ws = wb.active
    except Exception:
        raise BusinessException(400, "Excel 文件解析失败")

    # 扫描前 15 行，找到包含「学号」的表头行
    header_row_idx = None
    header_values = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        str_vals = [str(c).strip() if c is not None else "" for c in row]
        if any("学号" in v for v in str_vals):
            header_row_idx = i
            header_values = str_vals
            break

    if header_row_idx is None or header_values is None:
        raise BusinessException(400, "未找到表头行，请确认文件包含「学号」列")

    # 定位学号列和姓名列
    id_col = None
    name_col = None
    for idx, val in enumerate(header_values):
        if val == "学号":
            id_col = idx
        elif val == "姓名":
            name_col = idx

    if id_col is None:
        raise BusinessException(400, "未找到「学号」列")
    if name_col is None:
        raise BusinessException(400, "未找到「姓名」列")

    # 专业列 = 姓名列的下一列
    major_col = name_col + 1

    # 遍历数据行
    result = {"success_count": 0, "skip_count": 0, "fail_count": 0, "errors": []}

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        row_list = list(row)
        if id_col >= len(row_list) or name_col >= len(row_list):
            continue

        student_id = str(row_list[id_col]).strip() if row_list[id_col] is not None else ""
        name = str(row_list[name_col]).strip() if row_list[name_col] is not None else ""

        if not student_id or not name:
            continue

        # 提取专业（可能为 None）
        major = ""
        if major_col < len(row_list) and row_list[major_col] is not None:
            major = str(row_list[major_col]).strip()

        try:
            with db.begin_nested():
                student = db.query(User).filter(User.id == student_id).first()
                is_new = False
                if not student:
                    student = User(
                        id=student_id,
                        name=name,
                        hashed_password=get_password_hash(DEFAULT_STUDENT_PASSWORD),
                        role="student",
                        major=major,
                    )
                    db.add(student)
                    db.flush()
                    is_new = True
                else:
                    student.name = name
                    student.major = major

                # 指定了班级时，自动注册
                if target_class is not None:
                    existing_enrollment = db.query(StudentClassEnrollment).filter(
                        StudentClassEnrollment.user_id == student_id,
                        StudentClassEnrollment.class_id == target_class.id,
                    ).first()
                    if not existing_enrollment:
                        db.add(StudentClassEnrollment(
                            user_id=student_id, class_id=target_class.id))
                        db.flush()

                if is_new:
                    result["success_count"] += 1
                else:
                    result["skip_count"] += 1
        except (SQLAlchemyError, IntegrityError) as exc:
            result["fail_count"] += 1
            result["errors"].append({"row": row_idx, "reason": str(exc)[:120]})
            continue

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise BusinessException(500, "批量导入失败")
    except SQLAlchemyError:
        db.rollback()
        raise BusinessException(500, "批量导入失败")

    return result
